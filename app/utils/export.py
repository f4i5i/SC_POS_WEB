"""
Export utilities for generating Excel, CSV, and PDF exports
"""

from io import BytesIO
from datetime import datetime
import csv

# Try to import openpyxl for Excel export
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


def export_to_excel(data, columns, title="Report", sheet_name="Data"):
    """
    Export data to Excel format

    Args:
        data: List of dictionaries or list of lists containing the data
        columns: List of column headers or dict mapping keys to display names
        title: Report title for the header
        sheet_name: Name of the worksheet

    Returns:
        BytesIO object containing the Excel file
    """
    if not EXCEL_AVAILABLE:
        raise ImportError("openpyxl is required for Excel export")

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    title_font = Font(bold=True, size=14)
    date_font = Font(italic=True, size=10, color="666666")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal='center')

    # Date
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(columns))
    date_cell = ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    date_cell.font = date_font
    date_cell.alignment = Alignment(horizontal='center')

    # Headers
    header_row = 4
    if isinstance(columns, dict):
        headers = list(columns.values())
        keys = list(columns.keys())
    else:
        headers = columns
        keys = columns

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # Data rows
    for row_idx, row_data in enumerate(data, header_row + 1):
        for col_idx, key in enumerate(keys, 1):
            if isinstance(row_data, dict):
                value = row_data.get(key, '')
            else:
                value = row_data[col_idx - 1] if col_idx - 1 < len(row_data) else ''

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

            # Format numbers
            if isinstance(value, (int, float)):
                cell.alignment = Alignment(horizontal='right')
            else:
                cell.alignment = Alignment(horizontal='left')

    # Adjust column widths
    for col_idx in range(1, len(columns) + 1):
        column_letter = get_column_letter(col_idx)
        max_length = 0
        for cell in ws[column_letter]:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_to_csv(data, columns, include_header=True):
    """
    Export data to CSV format

    Args:
        data: List of dictionaries or list of lists containing the data
        columns: List of column headers or dict mapping keys to display names
        include_header: Whether to include header row

    Returns:
        BytesIO object containing the CSV file
    """
    output = BytesIO()

    if isinstance(columns, dict):
        headers = list(columns.values())
        keys = list(columns.keys())
    else:
        headers = columns
        keys = columns

    # Use text mode wrapper for csv
    import io
    text_output = io.StringIO()
    writer = csv.writer(text_output)

    if include_header:
        writer.writerow(headers)

    for row_data in data:
        if isinstance(row_data, dict):
            row = [row_data.get(key, '') for key in keys]
        else:
            row = row_data
        writer.writerow(row)

    # Convert to bytes
    output.write(text_output.getvalue().encode('utf-8-sig'))  # BOM for Excel compatibility
    output.seek(0)
    return output


def format_currency(value, symbol="Rs."):
    """Format a number as currency"""
    try:
        return f"{symbol} {float(value):,.2f}"
    except (TypeError, ValueError):
        return f"{symbol} 0.00"


def format_date(dt, format_str='%Y-%m-%d'):
    """Format a datetime object"""
    if dt:
        if isinstance(dt, str):
            return dt
        return dt.strftime(format_str)
    return ''


# Export templates for common report types
def export_sales_report(sales, format_type='excel'):
    """
    Export sales report data

    Args:
        sales: List of Sale objects or dicts
        format_type: 'excel' or 'csv'

    Returns:
        BytesIO object with the file
    """
    columns = {
        'invoice_number': 'Invoice #',
        'date': 'Date',
        'customer': 'Customer',
        'items': 'Items',
        'subtotal': 'Subtotal',
        'discount': 'Discount',
        'total': 'Total',
        'payment_method': 'Payment',
        'cashier': 'Cashier'
    }

    data = []
    for sale in sales:
        if hasattr(sale, '__dict__'):
            data.append({
                'invoice_number': sale.invoice_number or '',
                'date': format_date(sale.sale_date, '%Y-%m-%d %H:%M'),
                'customer': sale.customer.name if sale.customer else 'Walk-in',
                'items': len(sale.items) if hasattr(sale, 'items') else 0,
                'subtotal': float(sale.subtotal or 0),
                'discount': float(sale.discount_amount or 0),
                'total': float(sale.total or 0),
                'payment_method': (sale.payment_method or 'Cash').title(),
                'cashier': sale.user.full_name if sale.user else ''
            })
        else:
            data.append(sale)

    if format_type == 'excel':
        return export_to_excel(data, columns, title="Sales Report")
    else:
        return export_to_csv(data, columns)


def export_inventory_report(products, format_type='excel'):
    """
    Export inventory report data

    Args:
        products: List of Product objects or dicts
        format_type: 'excel' or 'csv'

    Returns:
        BytesIO object with the file
    """
    columns = {
        'code': 'Product Code',
        'name': 'Product Name',
        'category': 'Category',
        'brand': 'Brand',
        'quantity': 'Stock',
        'cost_price': 'Cost Price',
        'selling_price': 'Selling Price',
        'stock_value': 'Stock Value'
    }

    data = []
    for product in products:
        if hasattr(product, '__dict__'):
            data.append({
                'code': product.code or '',
                'name': product.name or '',
                'category': product.category.name if product.category else '',
                'brand': product.brand or '',
                'quantity': product.quantity or 0,
                'cost_price': float(product.cost_price or 0),
                'selling_price': float(product.selling_price or 0),
                'stock_value': float((product.cost_price or 0) * (product.quantity or 0))
            })
        else:
            data.append(product)

    if format_type == 'excel':
        return export_to_excel(data, columns, title="Inventory Report")
    else:
        return export_to_csv(data, columns)


def export_customer_report(customers, format_type='excel'):
    """
    Export customer report data

    Args:
        customers: List of Customer objects or dicts
        format_type: 'excel' or 'csv'

    Returns:
        BytesIO object with the file
    """
    columns = {
        'name': 'Customer Name',
        'phone': 'Phone',
        'email': 'Email',
        'total_purchases': 'Total Purchases',
        'loyalty_points': 'Loyalty Points',
        'loyalty_tier': 'Tier',
        'last_purchase': 'Last Purchase'
    }

    data = []
    for customer in customers:
        if hasattr(customer, '__dict__'):
            data.append({
                'name': customer.name or '',
                'phone': customer.phone or '',
                'email': customer.email or '',
                'total_purchases': float(customer.total_purchases or 0),
                'loyalty_points': customer.loyalty_points or 0,
                'loyalty_tier': customer.loyalty_tier or 'Bronze',
                'last_purchase': format_date(customer.last_purchase_date)
            })
        else:
            data.append(customer)

    if format_type == 'excel':
        return export_to_excel(data, columns, title="Customer Report")
    else:
        return export_to_csv(data, columns)
